"""
CAPA DE INFRAESTRUCTURA — Exportador a Excel (openpyxl).

Implementa el puerto PuertoExportadorPronostico. Produce libros .xlsx con
formato profesional y, como elemento de integridad de la tesis, el sistema
calcula el hash SHA-256 de cada archivo generado (función `sha256_de`).

Hojas del pronóstico (Módulo 2): Pronóstico, Resumen, Parámetros.
Hojas de la evaluación (Módulo 1): Resultados, Pruebas estadísticas.
"""

from __future__ import annotations

import hashlib
import io
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.aplicacion.puertos import PuertoExportadorPronostico
from src.dominio.entidades import Pronostico, SerieTemporal

AZUL_CABECERA = "1F4E5F"  # Azul marino institucional.
DORADO_GANADOR = "F4E5C2"  # Relleno suave para resaltar al ganador.


def sha256_de(contenido: bytes) -> str:
    """Huella criptográfica SHA-256 del archivo (integridad verificable)."""
    return hashlib.sha256(contenido).hexdigest()


class ExportadorExcel(PuertoExportadorPronostico):
    """Convierte resultados del dominio en libros Excel con estilo."""

    # ------------------------------------------------------------------ #
    # Estilos reutilizables                                              #
    # ------------------------------------------------------------------ #
    def _estilizar_cabecera(self, hoja, n_columnas: int) -> None:
        """Pinta la fila 1: fondo azul, letra blanca en negrita, centrada."""
        relleno = PatternFill("solid", fgColor=AZUL_CABECERA)
        fuente = Font(color="FFFFFF", bold=True)
        centrado = Alignment(horizontal="center", vertical="center")
        for col in range(1, n_columnas + 1):
            celda = hoja.cell(row=1, column=col)
            celda.fill = relleno
            celda.font = fuente
            celda.alignment = centrado

    def _bordear_y_ensanchar(self, hoja, n_filas: int, n_columnas: int) -> None:
        """Bordes finos en toda la tabla y anchos de columna legibles."""
        fino = Side(style="thin", color="B0B7BD")
        borde = Border(left=fino, right=fino, top=fino, bottom=fino)
        for fila in range(1, n_filas + 1):
            for col in range(1, n_columnas + 1):
                hoja.cell(row=fila, column=col).border = borde
        for col in range(1, n_columnas + 1):
            hoja.column_dimensions[get_column_letter(col)].width = 22
        hoja.freeze_panes = "A2"  # La cabecera queda fija al desplazarse.

    def _a_bytes(self, libro: Workbook) -> bytes:
        """Guarda el libro en memoria (sin tocar el disco) y devuelve sus bytes."""
        flujo = io.BytesIO()
        libro.save(flujo)
        return flujo.getvalue()

    # ------------------------------------------------------------------ #
    # Módulo 2: el Excel que descarga el dueño del retail                #
    # ------------------------------------------------------------------ #
    def exportar_pronostico(
        self,
        serie: SerieTemporal,
        pronostico: Pronostico,
        parametros_descripcion: List[tuple],
    ) -> bytes:
        libro = Workbook()

        # --- Hoja 1: Pronóstico (la tabla de fechas y unidades) ---------------
        hoja = libro.active
        hoja.title = "Pronóstico"
        hoja.append(["Fecha", "Pronóstico (unidades)", "Límite inferior", "Límite superior"])
        hay_limites = pronostico.limites_inferiores is not None
        for i, fecha in enumerate(pronostico.fechas):
            fila = [
                fecha,
                round(pronostico.valores[i], 2),
                round(pronostico.limites_inferiores[i], 2) if hay_limites else None,
                round(pronostico.limites_superiores[i], 2) if hay_limites else None,
            ]
            hoja.append(fila)
        for fila in hoja.iter_rows(min_row=2):
            fila[0].number_format = "dd/mm/yyyy"
            for celda in fila[1:]:
                celda.number_format = "#,##0.00"
        self._estilizar_cabecera(hoja, 4)
        self._bordear_y_ensanchar(hoja, hoja.max_row, 4)

        # --- Hoja 2: Resumen para el dueño -------------------------------------
        resumen = libro.create_sheet("Resumen")
        total = sum(pronostico.valores)
        resumen.append(["Indicador", "Valor"])
        resumen.append(["Producto / serie", serie.nombre])
        resumen.append(["Días pronosticados", len(pronostico.fechas)])
        resumen.append(["Total proyectado (unidades)", round(total, 2)])
        resumen.append(
            [
                "Promedio diario proyectado",
                round(total / max(1, len(pronostico.fechas)), 2),
            ]
        )
        resumen.append(["Generado con", "Prophet (Taylor & Letham, 2018)"])
        self._estilizar_cabecera(resumen, 2)
        self._bordear_y_ensanchar(resumen, resumen.max_row, 2)

        # --- Hoja 3: Parámetros (trazabilidad del experimento) ------------------
        params = libro.create_sheet("Parámetros")
        params.append(["Parámetro", "Valor"])
        for etiqueta, valor in parametros_descripcion:
            params.append([etiqueta, valor])
        self._estilizar_cabecera(params, 2)
        self._bordear_y_ensanchar(params, params.max_row, 2)

        return self._a_bytes(libro)

    # ------------------------------------------------------------------ #
    # Módulo 1: la evidencia del experimento para los anexos de la tesis #
    # ------------------------------------------------------------------ #
    def exportar_evaluacion(self, filas: List[dict], pruebas: List[dict]) -> bytes:
        libro = Workbook()

        hoja = libro.active
        hoja.title = "Resultados"
        cabeceras = list(filas[0].keys()) if filas else ["Sin resultados"]
        hoja.append(cabeceras)
        for fila in filas:
            hoja.append(list(fila.values()))
        # Resaltar en dorado las filas marcadas como ganadoras.
        if filas and "Ganador" in cabeceras:
            col_ganador = cabeceras.index("Ganador") + 1
            relleno = PatternFill("solid", fgColor=DORADO_GANADOR)
            for numero_fila in range(2, hoja.max_row + 1):
                if hoja.cell(row=numero_fila, column=col_ganador).value == "Sí":
                    for col in range(1, len(cabeceras) + 1):
                        hoja.cell(row=numero_fila, column=col).fill = relleno
        self._estilizar_cabecera(hoja, len(cabeceras))
        self._bordear_y_ensanchar(hoja, hoja.max_row, len(cabeceras))

        if pruebas:
            hoja2 = libro.create_sheet("Pruebas estadísticas")
            cabeceras2 = list(pruebas[0].keys())
            hoja2.append(cabeceras2)
            for fila in pruebas:
                hoja2.append(list(fila.values()))
            self._estilizar_cabecera(hoja2, len(cabeceras2))
            self._bordear_y_ensanchar(hoja2, hoja2.max_row, len(cabeceras2))

        return self._a_bytes(libro)
