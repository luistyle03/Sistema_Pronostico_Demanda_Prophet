"""[S15 · Iteración 6] Exportador Excel: la evidencia descargable."""

import io
from datetime import date, timedelta

from openpyxl import load_workbook

from src.aplicacion.parametros import ParametrosPronostico
from src.dominio.entidades import Pronostico, PuntoSerie, SerieTemporal
from src.infraestructura.modelos.utilidades import fechas_futuras
from src.infraestructura.persistencia.exportador_excel import ExportadorExcel, sha256_de


def test_exporta_un_xlsx_valido_y_trazable():
    inicio = date(2026, 1, 1)
    serie = SerieTemporal(
        "Gaseosa 500ml",
        [PuntoSerie(inicio + timedelta(days=i), 5.0) for i in range(30)],
    )
    pron = Pronostico("Prophet", fechas_futuras(serie.fechas()[-1], 7), [6.0] * 7)
    contenido = ExportadorExcel().exportar_pronostico(
        serie, pron, ParametrosPronostico().descripcion()
    )
    assert contenido[:2] == b"PK"  # firma de todo .xlsx (zip)
    libro = load_workbook(io.BytesIO(contenido))
    assert any(
        "metro" in nombre.lower() for nombre in libro.sheetnames
    )  # hoja Parámetros (sin tropezar con la tilde)
    assert len(sha256_de(contenido)) == 64  # huella de integridad
