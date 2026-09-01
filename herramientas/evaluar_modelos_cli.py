#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""evaluar_modelos_cli.py — Evaluación comparativa por consola.

Ejecuta EL MISMO núcleo que la aplicación web (src/), de modo que los números
que imprime son idénticos a los del Módulo 1. No reimplementa métricas ni
modelos: importa los adaptadores y el caso de uso EvaluadorDeModelos.

Uso:
    python herramientas/evaluar_modelos_cli.py favorita_50_series.csv \
        --fraccion-prueba 0.2 --salida resultados_evaluacion.xlsx
"""

from __future__ import annotations

import argparse
import math
import platform
import sys
from datetime import datetime
from pathlib import Path

RAIZ = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(RAIZ))

from openpyxl import Workbook

# Se reutiliza EXACTAMENTE el mismo núcleo del experimento principal.
from src.aplicacion.casos_uso.evaluar_modelos import EvaluadorDeModelos
from src.aplicacion.parametros import ParametrosPronostico
from src.infraestructura.estadistica.adaptador_scipy import AdaptadorPruebasScipy
from src.infraestructura.modelos.adaptador_arima import AdaptadorARIMA
from src.infraestructura.modelos.adaptador_holt_winters import AdaptadorHoltWinters
from src.infraestructura.modelos.adaptador_media_movil import AdaptadorMediaMovil
from src.infraestructura.modelos.adaptador_prophet import AdaptadorProphet
from src.infraestructura.modelos.adaptador_regresion_lineal import (
    AdaptadorRegresionLineal,
)
from src.infraestructura.persistencia.lector_archivos import LectorVentas

MODELOS = ["Prophet", "ARIMA", "Holt-Winters", "Promedio móvil", "Regresión lineal"]


def _construir_evaluador() -> EvaluadorDeModelos:
    """Los mismos 5 modelos del experimento principal (Prophet con feriados EC)."""
    modelos = [
        AdaptadorProphet(ParametrosPronostico(pais_feriados="EC")),
        AdaptadorARIMA(),
        AdaptadorHoltWinters(),
        AdaptadorMediaMovil(),
        AdaptadorRegresionLineal(),
    ]
    return EvaluadorDeModelos(modelos, AdaptadorPruebasScipy())


# === Comparación operativa: capacidades DOCUMENTADAS (no calculadas) =========
# Derivadas de los requisitos del retail y de la documentación oficial de cada
# modelo. Escala 2=Sí, 1=Parcial, 0=No. La fila "Velocidad" NO está aquí: se
# completa empíricamente con los tiempos de ejecución medidos por el software.
# La lista es BALANCEADA: incluye criterios donde Prophet NO gana (velocidad,
# simplicidad), para no sesgar el resultado a su favor.
def guardar_excel(lote, destino: Path, entrada: Path, fraccion: float) -> None:
    """Evidencia descargable: una fila por serie y modelo, más las pruebas."""
    wb = Workbook()

    # --- Hoja 1: ficha técnica de la corrida (trazabilidad de la medición) ---
    ficha = wb.active
    ficha.title = "Ficha técnica"
    ficha.append(["FICHA TÉCNICA DE LA CORRIDA"])
    ficha.append([])
    for campo, valor in [
        (
            "Instrumento",
            "Módulo de evaluación automatizada del desempeño del pronóstico",
        ),
        ("Investigador", "Pedro Alberto Luis Méndez"),
        ("Fecha y hora de la corrida", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
        (
            "Equipo",
            f"{platform.processor() or platform.machine()} · "
            f"{platform.system()} {platform.release()}",
        ),
        ("Versión de Python", platform.python_version()),
        ("Archivo de entrada", str(entrada)),
        ("Series evaluadas", len(lote.detalle_por_serie)),
        ("Series omitidas", len(lote.series_omitidas)),
        ("Modelos comparados", ", ".join(MODELOS)),
        (
            "Técnica de evaluación",
            "Out-of-sample con partición temporal de origen fijo (holdout)",
        ),
        ("Porción de prueba (holdout)", f"{fraccion:.0%} final de cada serie"),
        ("Historia mínima exigida", "365 días (regla RF03)"),
    ]:
        ficha.append([campo, valor])

    ws = wb.create_sheet("Resultados")
    ws.append(
        [
            "Serie",
            "Modelo",
            "RMSSE",
            "WAPE (%)",
            "MAE (unid.)",
            "Sesgo (unid.)",
            "RMSE",
            "MAPE (%)",
            "Tiempo (s)",
            "Ganador",
            "Error",
        ]
    )
    for evaluacion in lote.detalle_por_serie:
        ganador = evaluacion.ganador
        nombre_ganador = ganador.nombre_modelo if ganador else None
        for r in evaluacion.resultados:
            ws.append(
                [
                    evaluacion.nombre_serie,
                    r.nombre_modelo,
                    r.rmsse,
                    r.wape,
                    r.mae,
                    r.sesgo,
                    r.rmse,
                    r.mape,
                    r.segundos,
                    "Sí" if r.nombre_modelo == nombre_ganador else "No",
                    getattr(r, "mensaje_error", None),
                ]
            )
    ws2 = wb.create_sheet("Pruebas estadísticas")
    ws2.append(
        [
            "Comparación",
            "p-valor (t pareada)",
            "p-valor (Wilcoxon)",
            "d de Cohen",
            "N (pares)",
        ]
    )
    for p in lote.pruebas:
        ws2.append([p.comparacion, p.p_valor_t, p.p_valor_wilcoxon, p.d_cohen, p.n])
    wb.save(destino)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Evaluación comparativa (mismo núcleo que la app)."
    )
    parser.add_argument("entrada", type=Path, help="CSV/XLSX con fecha, serie/producto y ventas")
    parser.add_argument(
        "--fraccion-prueba",
        type=float,
        default=0.20,
        help="Porción final reservada para prueba (0.20 = 20 %%, plan de tesis)",
    )
    parser.add_argument("--salida", type=Path, default=Path("resultados_evaluacion.xlsx"))
    args = parser.parse_args()

    lector = LectorVentas()
    tabla = lector.leer(args.entrada.read_bytes(), str(args.entrada))
    series = lector.construir_series_lote(tabla)
    print(f"Dataset cargado: {len(series)} serie(s).")
    print(f"Holdout temporal: {args.fraccion_prueba:.0%} final de cada serie.\n")

    lote = _construir_evaluador().ejecutar_lote(series, args.fraccion_prueba)

    n = len(lote.detalle_por_serie)
    print("\n" + "=" * 86)
    print(f"RESULTADO SOBRE {n} SERIE(S) — métrica principal: RMSSE (mediana; menor mejor)")
    print("=" * 86)
    print(
        f"{'Modelo':<18}{'RMSSEmed':>10}{'RMSSEprom':>11}{'WAPEmed%':>10}"
        f"{'MAEprom':>9}{'Sesgo':>8}{'Gana':>7}{'Sup<1':>8}"
    )
    ganador = lote.ganador
    for r in lote.resumen_por_modelo:
        marca = "  <-- GANADOR" if ganador and r.nombre_modelo == ganador.nombre_modelo else ""
        print(
            f"{r.nombre_modelo:<18}{r.rmsse_mediana:>10.4f}{r.rmsse_promedio:>11.4f}"
            f"{r.wape_mediana:>10.1f}{r.mae_promedio:>9.1f}{r.sesgo_promedio:>8.1f}"
            f"{str(r.series_ganadas) + '/' + str(r.series_evaluadas):>7}"
            f"{str(r.series_supera_ingenuo) + '/' + str(r.series_evaluadas):>8}{marca}"
        )

    # Se reportan los estadisticos SIN emitir veredicto. La regla de decision
    # declarada a priori en la seccion 3.6 de la tesis exige la correccion de
    # Holm sobre la familia de cuatro contrastes y el criterio de tamano del
    # efecto; ambas cosas las aplica contraste_hipotesis.py, que es la unica
    # fuente estadistica confirmatoria del estudio. Emitir aqui un veredicto
    # con una regla distinta produciria dos metodologias para una misma corrida.
    print("\nPRUEBAS PAREADAS (Prophet frente a cada clasico, sobre RMSSE)")
    print("  Valores p SIN corregir por multiplicidad. El veredicto de las hipotesis")
    print("  HE2 a HE5 se emite unicamente con herramientas/contraste_hipotesis.py.")
    for p in lote.pruebas:
        print(
            f"  {p.comparacion:<30}p(t)={p.p_valor_t:.4f}  p(W)={p.p_valor_wilcoxon:.4f}  "
            f"|d|={p.d_cohen:.3f}"
        )

    if lote.series_omitidas:
        print(f"\nSeries omitidas por longitud insuficiente: {len(lote.series_omitidas)}")

    # La idoneidad operativa (HE7) NO la calcula este programa: es un analisis
    # documental que el investigador diligencia sobre la matriz del Instrumento B,
    # citando la fuente de cada calificacion. El unico insumo empirico que este
    # programa aporta a esa matriz son los tiempos medidos, que se imprimen para
    # que se apliquen los umbrales absolutos de la rubrica validada.
    print("\nTIEMPOS MEDIANOS POR SERIE (insumo empirico del criterio de velocidad)")
    for r in sorted(lote.resumen_por_modelo, key=lambda x: x.segundos_promedio or math.inf):
        t = r.segundos_promedio
        print(f"  {r.nombre_modelo:<20}{t:8.4f} s" if t is not None else f"  {r.nombre_modelo:<20}    n/d")
    print("  Los niveles del criterio se asignan con la rubrica del Anexo A.4 del")
    print("  expediente de validacion, no con un ordenamiento relativo.")

    guardar_excel(lote, args.salida, args.entrada, args.fraccion_prueba)
    print(f"\nEvidencia guardada en: {args.salida}")


if __name__ == "__main__":
    main()
